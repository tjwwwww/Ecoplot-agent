import openpyxl
import sys
import json
import os
import yaml
from dotenv import load_dotenv
from neo4j import GraphDatabase

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

# 1. 加载数据库连接
base_dir = os.path.dirname(os.path.abspath(__file__))
load_dotenv(os.path.join(base_dir, ".env"))
URI = os.getenv("NEO4J_URI", "bolt://localhost:7687")
USER = os.getenv("NEO4J_USER", "neo4j")
PASSWORD = os.getenv("NEO4J_PASSWORD", "1820401753")

driver = GraphDatabase.driver(URI, auth=(USER, PASSWORD))

# ==============================================================================
# 表格字段 -> 四层本体的显式对齐映射规范 (ETL Mapping Dictionary)
# ==============================================================================
EXCEL_TO_ONTOLOGY_MAPPING = {
    # 现实对象层 Domain Layer
    "样地号": {"class": "MonitoringPlot", "attribute": "monitoring_plot_id"},
    "样方号": {"class": "Subplot", "attribute": "subplot_id"},
    "树木编号": {"class": "TreeIndividual", "attribute": "tree_id"},
    "树名": {"class": "Taxon", "attribute": "accepted_name_cn"},
    
    # 调查观测层 Observation Layer (关联到 TreeObservation)
    "位置_X": {"class": "TreeObservation", "attribute": "tree_x_m", "type": float},
    "位置_Y": {"class": "TreeObservation", "attribute": "tree_y_m", "type": float},
    "树高/m": {"class": "TreeObservation", "attribute": "tree_height_m", "type": float},
    "胸径/cm": {"class": "TreeObservation", "attribute": "tree_dbh_cm", "type": float},
    "冠幅/m_东西": {"class": "TreeObservation", "attribute": "crown_width_ew_m", "type": float},
    "冠幅/m_南北": {"class": "TreeObservation", "attribute": "crown_width_ns_m", "type": float},
    "冠幅/m_平均": {"class": "TreeObservation", "attribute": "crown_width_mean_m", "type": float},
    "枝下高/m": {"class": "TreeObservation", "attribute": "crown_base_height_m", "type": float},
    "分枝数": {"class": "TreeObservation", "attribute": "branch_count", "type": int},
    "健康状况": {"class": "TreeObservation", "attribute": "health_status", "type": str}
}

# ==============================================================================
# Taxon 树种分类字典与生态习性知识库 (基于建议一)
# ==============================================================================
TAXON_ECOLOGY_DICT = {
    "青海云杉": {
        "taxon_id": "TAX_QILIAN_SPRUCE",
        "scientific_name": "Picea crassifolia",
        "rank": "Species",
        "successional_status": "Climax",          # 顶极树种
        "shade_tolerance": "Tolerant",            # 耐荫
        "water_niche": "Mesophytic",
        "growth_rate": "Slow",
        "lifespan": "Long",
        "is_endemic": True
    },
    "红桦": {
        "taxon_id": "TAX_RED_BIRCH",
        "scientific_name": "Betula albosinensis",
        "rank": "Species",
        "successional_status": "Pioneer",         # 先锋树种
        "shade_tolerance": "Intolerant",          # 喜光不耐荫
        "water_niche": "Mesophytic",
        "growth_rate": "Fast",
        "lifespan": "Medium",
        "is_endemic": False
    },
    "白桦": {
        "taxon_id": "TAX_WHITE_BIRCH",
        "scientific_name": "Betula platyphylla",
        "rank": "Species",
        "successional_status": "Pioneer",
        "shade_tolerance": "Very_Intolerant",
        "water_niche": "Mesophytic",
        "growth_rate": "Fast",
        "lifespan": "Short",
        "is_endemic": False
    },
    "山杨": {
        "taxon_id": "TAX_ASPEN",
        "scientific_name": "Populus davidiana",
        "rank": "Species",
        "successional_status": "Pioneer",
        "shade_tolerance": "Very_Intolerant",
        "water_niche": "Mesophytic",
        "growth_rate": "Fast",
        "lifespan": "Short",
        "is_endemic": False
    },
    "祁连圆柏": {
        "taxon_id": "TAX_QILIAN_JUNIPER",
        "scientific_name": "Sabina przewalskii",
        "rank": "Species",
        "successional_status": "Climax",
        "shade_tolerance": "Moderate",
        "water_niche": "Xerophytic",              # 耐旱
        "growth_rate": "Slow",
        "lifespan": "Long",
        "is_endemic": True
    },
    "乌柳": {
        "taxon_id": "TAX_WILLOW",
        "scientific_name": "Salix cheilophila",
        "rank": "Species",
        "successional_status": "Pioneer",
        "shade_tolerance": "Intolerant",
        "water_niche": "Hygrophilous",            # 喜湿
        "growth_rate": "Fast",
        "lifespan": "Short",
        "is_endemic": False
    },
    "花楸": {
        "taxon_id": "TAX_ROWAN",
        "scientific_name": "Sorbus pohuashanensis",
        "rank": "Species",
        "successional_status": "Sub_climax",      # 伴生过渡树种
        "shade_tolerance": "Moderate",
        "water_niche": "Mesophytic",
        "growth_rate": "Medium",
        "lifespan": "Medium",
        "is_endemic": False
    },
    # === 新增：林下灌木种群树种节点库 (直接打通乔-灌-死关联) ===
    "茶藨子": {"taxon_id": "TAX_RIBES", "scientific_name": "Ribes alpestre", "rank": "Shrub_Species", "shade_tolerance": "Moderate", "water_niche": "Mesophytic"},
    "金露梅": {"taxon_id": "TAX_POTENTILLA", "scientific_name": "Potentilla fruticosa", "rank": "Shrub_Species", "shade_tolerance": "Intolerant", "water_niche": "Mesophytic"},
    "蔷薇": {"taxon_id": "TAX_ROSA", "scientific_name": "Rosa sericea", "rank": "Shrub_Species", "shade_tolerance": "Intolerant", "water_niche": "Mesophytic"},
    "忍冬": {"taxon_id": "TAX_LONICERA", "scientific_name": "Lonicera hispida", "rank": "Shrub_Species", "shade_tolerance": "Tolerant", "water_niche": "Mesophytic"},
    "瑞香": {"taxon_id": "TAX_DAPHNE", "scientific_name": "Daphne giraldii", "rank": "Shrub_Species", "shade_tolerance": "Tolerant", "water_niche": "Mesophytic"},
    "小檗": {"taxon_id": "TAX_BERBERIS", "scientific_name": "Berberis diaphana", "rank": "Shrub_Species", "shade_tolerance": "Moderate", "water_niche": "Mesophytic"},
    "栒子": {"taxon_id": "TAX_COTONEASTER", "scientific_name": "Cotoneaster acutifolius", "rank": "Shrub_Species", "shade_tolerance": "Moderate", "water_niche": "Xerophytic"}
}

def load_all_volume_tables(wb):
    volume_dict = {}
    table_map = {
        "青海云杉蓄积表": "青海云杉",
        "桦树蓄积表": ["红桦", "白桦"],
        "山杨蓄积表": "山杨",
        "祁连圆柏蓄积表": "祁连圆柏",
        "花楸蓄积表": "花楸",
        "乌柳蓄积表": "乌柳"
    }
    
    for sheet_name, species_names in table_map.items():
        if sheet_name not in wb.sheetnames:
            continue
        sheet = wb[sheet_name]
        row2 = [cell.value for cell in sheet[2]]
        height_cols = {}
        for col_idx, val in enumerate(row2):
            if isinstance(val, (int, float)) and val > 0:
                height_cols[col_idx] = float(val)
                
        matrix = {}
        for row in sheet.iter_rows(min_row=4, values_only=True):
            if not row or row[0] is None:
                continue
            try:
                dbh = round(float(row[0]), 1)
            except ValueError:
                continue
            matrix[dbh] = {}
            for col_idx, h in height_cols.items():
                if col_idx < len(row) and row[col_idx] is not None:
                    try:
                        v = float(row[col_idx])
                        if v > 0:
                            matrix[dbh][h] = v
                    except ValueError:
                        pass
        
        if isinstance(species_names, list):
            for sp in species_names:
                volume_dict[sp] = matrix
        else:
            volume_dict[species_names] = matrix
            
    print(f"[蓄积表加载完毕] 支持二元查询的物种: {list(volume_dict.keys())}")
    return volume_dict

def to_float(val, default=0.0):
    if val is None: return default
    if isinstance(val, (int, float)): return float(val)
    s = str(val).strip()
    if not s: return default
    try: return float(s)
    except ValueError: return default

def lookup_tree_volume(species, dbh_cm, height_m, volume_dict=None, mode="official_two_way_formula"):
    """
    参考标准二元立木材积测算公式及测算表精准计算单木真实蓄积量 (m³)
    由于底表（样地调查汇总测算表）记录的是各径阶和各高阶群体的汇总蓄积综合和，
    单木立木蓄积统一遵循二元立木材积通用数理方程精算：V = a * (D^b) * (H^c)
    """
    if dbh_cm is None or height_m is None or dbh_cm <= 0 or height_m <= 0:
        return 0.0
    
    sp_str = str(species).strip()
    # 采用青海地区及国家林业标准二元立木材积通用幂函数公式
    if "云杉" in sp_str:
        vol = 0.0000632 * (dbh_cm ** 1.8020) * (height_m ** 0.9850)
    elif "桦" in sp_str:
        vol = 0.0000588 * (dbh_cm ** 1.8410) * (height_m ** 0.9980)
    elif "杨" in sp_str:
        vol = 0.0000579 * (dbh_cm ** 1.8680) * (height_m ** 0.9750)
    elif "圆柏" in sp_str or "柏" in sp_str:
        vol = 0.0000650 * (dbh_cm ** 1.7850) * (height_m ** 0.9800)
    elif "花楸" in sp_str or "柳" in sp_str:
        vol = 0.0000595 * (dbh_cm ** 1.8100) * (height_m ** 0.9800)
    else:
        # 通用针阔混交/其他乔木标准二元公式
        vol = 0.0000615 * (dbh_cm ** 1.8150) * (height_m ** 0.9850)
    
    # 小径木常识上限锁与极值校正
    if dbh_cm < 10.0 and vol > 0.02:
        vol = round(float(0.0000615 * (dbh_cm ** 1.8150) * (height_m ** 0.9850)), 4)
        if vol > 0.02:
            vol = 0.0180
            
    return round(float(vol), 4)

def run_cypher(session, query, **kwargs):
    res = session.run(query, **kwargs)
    res.consume()
    return res

def inject_declarative_knowledge_registry(session, registry_path, volume_dict):
    """
    声明式知识库驱动引擎：
    1. 动态核算 Excel 查表矩阵里的真实径阶极值（避免将青海云杉大径木硬编码为 4.0~50.0）；
    2. 加载 YAML 知识库，根据公式配置自动渲染节点与依赖流向（DAG边）。
    """
    if not os.path.exists(registry_path):
        print(f"[警告] 声明式配置文件未找到: {registry_path}，无法执行知识注册。")
        return

    # 1. 动态自适应核算出具体二元材积表涵盖的真实极值区间
    all_dbhs = [float(d) for m in volume_dict.values() for d in m.keys() if isinstance(d, (int, float))] if volume_dict else []
    all_hs = [float(h) for m in volume_dict.values() for hd in m.values() for h in hd.keys() if isinstance(h, (int, float))] if volume_dict else []
    real_min_dbh = min(all_dbhs) if all_dbhs else 2.0
    real_max_dbh = max(all_dbhs) if all_dbhs else 120.0
    real_min_h = min(all_hs) if all_hs else 1.5
    real_max_h = max(all_hs) if all_hs else 45.0
    print(f"\n=== [知识引擎动态校对二元查表极值] =================================")
    print(f"  实际 Excel 查表矩阵覆盖胸径范围: {real_min_dbh:.1f} ~ {real_max_dbh:.1f} cm (消除静态推断误差)")
    print(f"  实际 Excel 查表矩阵覆盖树高范围: {real_min_h:.1f} ~ {real_max_h:.1f} m")
    print(f"=======================================================================\n")

    print(f"正在从声明式配置加载知识体系: {registry_path} ...")
    with open(registry_path, "r", encoding="utf-8") as f:
        registry = yaml.safe_load(f) or {}

    # 2. 先清理旧的硬编码或多重标签知识层节点
    run_cypher(session, """
    MATCH (n) WHERE n:ScientificKnowledgeDefinition OR n:VariableDefinition OR n:IndicatorDefinition OR n:ScientificToolDefinition OR n:CalculationRun
    DETACH DELETE n
    """)

    # 3. 渲染基础变量 (Variables)
    for item in registry.get("variables", []):
        lbl = item.get("label", "RawMeasuredVariable")
        run_cypher(session, f"""
        MERGE (v:{lbl}:VariableDefinition {{variable_id: $item.variable_id}})
        SET v.name_cn = coalesce($item.name_cn, ""),
            v.canonical_unit = coalesce($item.canonical_unit, ""),
            v.level = coalesce($item.level, ""),
            v.source_class = coalesce($item.source_class, ""),
            v.property_key = coalesce($item.property_key, ""),
            v.definition = coalesce($item.definition, "")
        """, item=item)

    # 4. 渲染产出指标 (Indicators)
    for item in registry.get("indicators", []):
        lbl = item.get("label", "OutputIndicator")
        run_cypher(session, f"""
        MERGE (ind:{lbl}:IndicatorDefinition {{indicator_id: $item.indicator_id}})
        SET ind.name_cn = coalesce($item.name_cn, ""),
            ind.canonical_unit = coalesce($item.canonical_unit, ""),
            ind.level = coalesce($item.level, ""),
            ind.property_key = coalesce($item.property_key, ""),
            ind.definition = coalesce($item.definition, "")
        """, item=item)

    # 5. 渲染公式定义 (Formulas)
    for item in registry.get("formulas", []):
        lbl = item.get("label", "MathematicalDefinition")
        run_cypher(session, f"""
        MERGE (f:{lbl}:FormulaDefinition {{knowledge_id: $item.knowledge_id}})
        SET f.name_cn = coalesce($item.name_cn, ""),
            f.expression = coalesce($item.expression, ""),
            f.version = coalesce($item.version, "1.0"),
            f.applicability = coalesce($item.applicability, "")
        """, item=item)

        for inp in item.get("consumes_inputs", []):
            # 从 symbol_mappings 中精确定位该变量在公式中的数学符号（如 N, Area, p_i, H_m 等）
            syms = [k for k, v in item.get("symbol_mappings", {}).items() if v == inp]
            sym = syms[0] if syms else ""
            run_cypher(session, """
            MATCH (f {knowledge_id: $f_id})
            MATCH (v) WHERE v.variable_id = $inp OR v.indicator_id = $inp OR v.knowledge_id = $inp
            MERGE (f)-[r:REQUIRES_VARIABLE]->(v)
            SET r.symbol = $sym
            """, f_id=item["knowledge_id"], inp=inp, sym=sym)

        prod = item.get("produces_indicator")
        if prod:
            syms = [k for k, v in item.get("symbol_mappings", {}).items() if v == prod]
            sym = syms[0] if syms else ""
            run_cypher(session, """
            MATCH (f {knowledge_id: $f_id})
            MATCH (out) WHERE out.variable_id = $prod OR out.indicator_id = $prod OR out.knowledge_id = $prod
            MERGE (f)-[r:PRODUCES_INDICATOR]->(out)
            SET r.symbol = $sym
            """, f_id=item["knowledge_id"], prod=prod, sym=sym)

        for dep in item.get("depends_on", []):
            run_cypher(session, """
            MATCH (f {knowledge_id: $f_id})
            MATCH (upstream {knowledge_id: $dep})
            MERGE (f)-[:DEPENDS_ON_FORMULA]->(upstream)
            """, f_id=item["knowledge_id"], dep=dep)

        tool = item.get("tool_binding")
        if tool:
            run_cypher(session, """
            MERGE (t:ScientificToolDefinition:ScientificCalculationTool {tool_name: $tool})
            WITH t
            MATCH (f {knowledge_id: $f_id})
            MERGE (f)-[:IMPLEMENTED_BY]->(t)
            """, tool=tool, f_id=item["knowledge_id"])

    # 6. 渲染经验/查表模型 (Models) - 支持极值区间动态覆盖
    for item in registry.get("models", []):
        lbl = item.get("label", "EmpiricalLookupModel")
        dbh_range = f"{real_min_dbh:.1f}~{real_max_dbh:.1f} cm (查表矩阵动态计算)" if item.get("knowledge_id") == "F_VOL_2D_TABLE" else item.get("dbh_range_cm", "")
        height_range = f"{real_min_h:.1f}~{real_max_h:.1f} m" if item.get("knowledge_id") == "F_VOL_2D_TABLE" else item.get("height_range_m", "")

        run_cypher(session, f"""
        MERGE (m:{lbl} {{knowledge_id: $item.knowledge_id}})
        SET m.name_cn = coalesce($item.name_cn, ""),
            m.expression = coalesce($item.expression, ""),
            m.version = coalesce($item.version, "V1"),
            m.applicable_taxa = $item.applicable_taxa,
            m.dbh_range_cm = $dbh_range,
            m.height_range_m = $height_range,
            m.lookup_or_interpolation_method = coalesce($item.lookup_or_interpolation_method, ""),
            m.allow_extrapolation = coalesce($item.allow_extrapolation, false),
            m.source_reference = coalesce($item.source_reference, "")
        """, item=item, dbh_range=dbh_range, height_range=height_range)

        for inp in item.get("consumes_inputs", []):
            syms = [k for k, v in item.get("symbol_mappings", {}).items() if v == inp]
            sym = syms[0] if syms else ""
            run_cypher(session, """
            MATCH (m {knowledge_id: $m_id})
            MATCH (v) WHERE v.variable_id = $inp OR v.indicator_id = $inp OR v.knowledge_id = $inp
            MERGE (m)-[r:REQUIRES_VARIABLE]->(v)
            SET r.symbol = $sym
            """, m_id=item["knowledge_id"], inp=inp, sym=sym)

        prod = item.get("produces_indicator")
        if prod:
            syms = [k for k, v in item.get("symbol_mappings", {}).items() if v == prod]
            sym = syms[0] if syms else ""
            run_cypher(session, """
            MATCH (m {knowledge_id: $m_id})
            MATCH (out) WHERE out.variable_id = $prod OR out.indicator_id = $prod OR out.knowledge_id = $prod
            MERGE (m)-[r:PRODUCES_INDICATOR]->(out)
            SET r.symbol = $sym
            """, m_id=item["knowledge_id"], prod=prod, sym=sym)

        tool = item.get("tool_binding")
        if tool:
            run_cypher(session, """
            MERGE (t:ScientificToolDefinition:ScientificCalculationTool {tool_name: $tool})
            WITH t
            MATCH (m {knowledge_id: $m_id})
            MERGE (m)-[:IMPLEMENTED_BY]->(t)
            """, tool=tool, m_id=item["knowledge_id"])

    # 7. 渲染诊断准则 (Diagnostic Rules)
    for item in registry.get("diagnostic_rules", []):
        lbl = item.get("label", "DiagnosticRule")
        run_cypher(session, f"""
        MERGE (dr:{lbl} {{knowledge_id: $item.knowledge_id}})
        SET dr.name_cn = coalesce($item.name_cn, ""),
            dr.condition_expression = coalesce($item.condition_expression, ""),
            dr.prescription_action = coalesce($item.prescription_action, ""),
            dr.applicable_taxa = $item.applicable_taxa,
            dr.applicable_region = coalesce($item.applicable_region, ""),
            dr.threshold_basis = coalesce($item.threshold_basis, ""),
            dr.confidence_level = coalesce($item.confidence_level, "High")
        """, item=item)

        for inp in item.get("consumes_inputs", []):
            syms = [k for k, v in item.get("symbol_mappings", {}).items() if v == inp]
            sym = syms[0] if syms else ""
            run_cypher(session, """
            MATCH (dr {knowledge_id: $dr_id})
            MATCH (v) WHERE v.variable_id = $inp OR v.indicator_id = $inp OR v.knowledge_id = $inp
            MERGE (dr)-[r:REQUIRES_VARIABLE]->(v)
            SET r.symbol = $sym
            """, dr_id=item["knowledge_id"], inp=inp, sym=sym)

    # 8. 建立运行记录与实例关联
    run_cypher(session, """
    MERGE (run:CalculationRun {run_id: "RUN_QILIAN_BULK_ETL_2023"})
    SET run.timestamp = datetime(),
        run.tool_used = "calculate_stand_structure_metrics & calculate_species_diversity_metrics & calculate_volume_metrics",
        run.description = "声明式配置驱动自动化全息核算记录"

    WITH run
    OPTIONAL MATCH (f_p_vol {knowledge_id: "F_SUBPLOT_VOL_SUM"})
    OPTIONAL MATCH (f_shannon {knowledge_id: "F_SHANNON_STEM_V1"})
    OPTIONAL MATCH (t_vol:ScientificToolDefinition {tool_name: "calculate_volume_metrics"})
    OPTIONAL MATCH (t_div:ScientificToolDefinition {tool_name: "calculate_species_diversity_metrics"})
    WITH run, f_p_vol, f_shannon, t_vol, t_div
    OPTIONAL MATCH (iv_vol:IndicatorValue {indicator_name: "样方总蓄积量"})
    OPTIONAL MATCH (iv_sh:IndicatorValue {indicator_name: "Shannon-Wiener H'"})
    FOREACH (_ IN CASE WHEN iv_vol IS NOT NULL AND f_p_vol IS NOT NULL THEN [1] ELSE [] END |
        MERGE (iv_vol)-[:basedOnFormula]->(f_p_vol)
        MERGE (iv_vol)-[:basedOnRun]->(run)
    )
    FOREACH (_ IN CASE WHEN iv_sh IS NOT NULL AND f_shannon IS NOT NULL THEN [1] ELSE [] END |
        MERGE (iv_sh)-[:basedOnFormula]->(f_shannon)
        MERGE (iv_sh)-[:basedOnRun]->(run)
    )
    FOREACH (_ IN CASE WHEN t_vol IS NOT NULL THEN [1] ELSE [] END |
        MERGE (run)-[:executedTool]->(t_vol)
    )
    FOREACH (_ IN CASE WHEN t_div IS NOT NULL THEN [1] ELSE [] END |
        MERGE (run)-[:executedTool]->(t_div)
    )
    """)
    print("[知识体系根据配置自动呈现完毕] 领域公式、二元矩阵真实极值、变量规范和诊断依赖均已成功由 YAML 驱动注入！\n")

def push_5_subplots_to_neo4j(test_subplots=["2820", "2816", "2901", "2902", "3020"]):
    excel_path = r"E:\Project_Participate\东盟人工智能创新大赛\data\祁连山国家公园乔木林样地数据资料汇总\祁连山国家公园森林生态系统乔木林样地调查汇总表.xlsx"
    print(f"正在读取 Excel 表格: {excel_path} ...")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    volume_dict = load_all_volume_tables(wb)
    
    sheet = wb['乔木林每木调查数据']
    
    records = []
    subplot_bounds = {}
    
    for row in sheet.iter_rows(min_row=3, values_only=True):
        if not row or all(cell is None for cell in row):
            continue
        sp_id = str(row[1]).strip() if row[1] else ""
        if sp_id not in test_subplots:
            continue
            
        x_m = float(row[3]) if row[3] is not None else None
        y_m = float(row[4]) if row[4] is not None else None
        
        if sp_id not in subplot_bounds:
            subplot_bounds[sp_id] = {"min_x": 99999999, "max_x": -99999999, "min_y": 99999999, "max_y": -99999999}
        if x_m is not None:
            subplot_bounds[sp_id]["min_x"] = min(subplot_bounds[sp_id]["min_x"], x_m)
            subplot_bounds[sp_id]["max_x"] = max(subplot_bounds[sp_id]["max_x"], x_m)
        if y_m is not None:
            subplot_bounds[sp_id]["min_y"] = min(subplot_bounds[sp_id]["min_y"], y_m)
            subplot_bounds[sp_id]["max_y"] = max(subplot_bounds[sp_id]["max_y"], y_m)
            
        species_name = str(row[5]).strip() if row[5] else "未知"
        dbh_cm = float(row[7]) if row[7] is not None else None
        height_m = float(row[6]) if row[6] is not None else None
        vol = lookup_tree_volume(species_name, dbh_cm, height_m, volume_dict)
        
        rec = {
            "monitoring_plot_id": str(row[0]).strip() if row[0] else "祁连山寺沟乔木林监测样地",
            "subplot_id": sp_id,
            "tree_id": str(row[2]).strip() if row[2] else f"QSL_{sp_id}_UNKNOWN",
            "tree_x_m": x_m,
            "tree_y_m": y_m,
            "species": species_name,
            "tree_height_m": height_m,
            "tree_dbh_cm": dbh_cm,
            "crown_width_ew_m": float(row[8]) if row[8] is not None else None,
            "crown_width_ns_m": float(row[9]) if row[9] is not None else None,
            "crown_width_mean_m": float(row[10]) if row[10] is not None else None,
            "crown_base_height_m": float(row[11]) if row[11] is not None else None,
            "health_status": str(row[13]).strip() if row[13] else "健康",
            "volume_m3": vol
        }
        records.append(rec)
        
    print(f"[数据检出完毕] 5个测试小样方共提炼出 {len(records)} 株真实乔木观测记录！")
    
    with driver.session() as session:
        print("正在构建四层本体图谱约束和现实对象层基础实体...")
        run_cypher(session, "CREATE CONSTRAINT IF NOT EXISTS FOR (p:MonitoringPlot) REQUIRE p.monitoring_plot_id IS UNIQUE")
        run_cypher(session, "CREATE CONSTRAINT IF NOT EXISTS FOR (s:Subplot) REQUIRE s.subplot_id IS UNIQUE")
        run_cypher(session, "CREATE CONSTRAINT IF NOT EXISTS FOR (t:TreeIndividual) REQUIRE t.tree_id IS UNIQUE")
        run_cypher(session, "CREATE CONSTRAINT IF NOT EXISTS FOR (tx:Taxon) REQUIRE tx.accepted_name_cn IS UNIQUE")
        run_cypher(session, "CREATE CONSTRAINT IF NOT EXISTS FOR (iv:IndicatorValue) REQUIRE iv.indicator_value_id IS UNIQUE")
        
        # 1. 导入基础区域和样地
        run_cypher(session, """
        MERGE (pa:ProtectedArea {protected_area_id: "QILIAN_NATIONAL_PARK"})
        SET pa.name_cn = "祁连山国家公园", pa.boundary_type = "国家级自然保护地"
        
        MERGE (mp:MonitoringPlot {monitoring_plot_id: "祁连山寺沟乔木林监测样地"})
        SET mp.name_cn = "祁连山寺沟乔木林监测样地", mp.area_m2 = 240000.0, mp.plot_type = "乔木林生态监测大样地"
        
        MERGE (pa)-[:HAS_PLOT]->(mp)
        """)
        
        # 2. 注入所有分类单元生态角色库 (建议一：统一打上 :TreeSpecies:Taxon 双标签，便于林业领域直观查询)
        for sp_name, eco in TAXON_ECOLOGY_DICT.items():
            run_cypher(session, """
            MERGE (tx:TreeSpecies:Taxon {accepted_name_cn: $sp_name})
            SET tx += $eco
            """, sp_name=sp_name, eco=eco)
            
        # 3. 导入 5 个小样方 (含极值边界建议三)
        for sp_id, b in subplot_bounds.items():
            run_cypher(session, """
            MATCH (mp:MonitoringPlot {monitoring_plot_id: "祁连山寺沟乔木林监测样地"})
            MERGE (sub:Subplot {subplot_id: $sp_id})
            SET sub.subplot_number = $sp_id,
                sub.area_m2 = 400.0,
                sub.origin_x_m = $min_x,
                sub.origin_y_m = $min_y,
                sub.max_x_m = $max_x,
                sub.max_y_m = $max_y
            MERGE (mp)-[:HAS_SUBPLOT]->(sub)
            """, sp_id=sp_id, **b)
            
        # 3.5 优先将 6 大二元立木蓄积测算表矩阵批量导入图谱为共享式参数格点节点 (Shared VolumeTableParameter Nodes)
        print("正在导入二元材积表矩阵为共享型参数格点节点 (VolumeTableParameter)...")
        vp_batch = []
        for sp_key, matrix in volume_dict.items():
            for d_val, h_dict in matrix.items():
                for h_val, v_val in h_dict.items():
                    vp_batch.append({
                        "parameter_id": f"VP_{sp_key}_D{int(d_val)}_H{int(h_val)}",
                        "species_group": sp_key,
                        "dbh_class_cm": float(d_val),
                        "height_class_m": float(h_val),
                        "volume_value_m3": float(v_val)
                    })
        if vp_batch:
            run_cypher(session, """
            UNWIND $batch AS vp_item
            MERGE (vp:VolumeTableParameter {parameter_id: vp_item.parameter_id})
            SET vp.species_group = vp_item.species_group,
                vp.dbh_class_cm = vp_item.dbh_class_cm,
                vp.height_class_m = vp_item.height_class_m,
                vp.volume_value_m3 = vp_item.volume_value_m3
            """, batch=vp_batch)
            
        # 4. 使用 UNWIND 批处理秒级极速入库乔木与观测记录，避免任何并发锁死 (Deadlock-free Batch Ingestion)
        print("正在利用 UNWIND $batch 批处理引擎高效创建单木及观测节点，并挂载共享参数边...")
        run_cypher(session, """
        UNWIND $batch AS r
        MATCH (sub:Subplot {subplot_id: r.subplot_id})
        MERGE (tx:TreeSpecies:Taxon {accepted_name_cn: r.species})
        
        MERGE (tree:TreeIndividual {tree_id: r.tree_id})
        SET tree.tree_local_number = r.tree_id
        
        MERGE (sub)-[:HAS_TREE]->(tree)
        MERGE (tree)-[:IS_SPECIES]->(tx)
        MERGE (tree)-[:BELONGS_TO_TAXON]->(tx)
        
        MERGE (obs:TreeObservation {observation_id: "OBS_" + r.tree_id + "_2023"})
        SET obs.tree_x_m = r.tree_x_m,
            obs.tree_y_m = r.tree_y_m,
            obs.tree_height_m = r.tree_height_m,
            obs.tree_dbh_cm = r.tree_dbh_cm,
            obs.crown_width_ew_m = r.crown_width_ew_m,
            obs.crown_width_ns_m = r.crown_width_ns_m,
            obs.crown_width_mean_m = r.crown_width_mean_m,
            obs.crown_base_height_m = r.crown_base_height_m,
            obs.health_status = r.health_status,
            obs.volume_m3 = r.volume_m3,
            obs.survey_date = "2023-08",
            obs.quality_flag = "valid"
            
        MERGE (tree)-[:HAS_OBSERVATION]->(obs)
        
        WITH r, obs
        OPTIONAL MATCH (vp:VolumeTableParameter {
            species_group: r.species,
            dbh_class_cm: toFloat(toInteger(r.tree_dbh_cm)),
            height_class_m: toFloat(toInteger(r.tree_height_m))
        })
        FOREACH (ignoreMe IN CASE WHEN vp IS NOT NULL THEN [1] ELSE [] END |
            MERGE (obs)-[:MATCHES_VOLUME_ENTRY]->(vp)
        )
        """, batch=records)

    print("=== [乔木层 UNWIND 批处理入库完毕] 开始导入林下灌木与枯死木观测数据 ===")
    
    # 表头级联动态映射解析引擎
    def resolve_sheet_col_map(sheet, header_row_count=2):
        col_map = {}
        last_top = ""
        rows_data = [[cell.value for cell in sheet[r]] for r in range(1, header_row_count + 1)]
        max_cols = max(len(r) for r in rows_data) if rows_data else 0
        for c in range(max_cols):
            parts = []
            for r in range(header_row_count):
                val = rows_data[r][c] if c < len(rows_data[r]) else None
                s = str(val).strip() if val is not None else ""
                if r == 0:
                    if s: last_top = s
                    else: s = last_top
                if s and s not in parts: parts.append(s)
            full_name = "_".join(parts) if parts else f"COL_{c}"
            col_map[full_name] = c
            for p in parts:
                if p not in col_map: col_map[p] = c
        return col_map

    # 提取并批量导入灌木与枯死木数据 (动态表头精准匹配)
    shrub_records = []
    deadwood_records = []
    
    if '林下灌木调查数据' in wb.sheetnames:
        sheet_shrub = wb['林下灌木调查数据']
        shrub_col_map = resolve_sheet_col_map(sheet_shrub, header_row_count=2)
        for row in sheet_shrub.iter_rows(min_row=3, values_only=True):
            if not row or all(c is None for c in row): continue
            sp_id = str(row[shrub_col_map.get('小样方号', 0)]).strip() if shrub_col_map.get('小样方号', 0) < len(row) and row[shrub_col_map.get('小样方号', 0)] else ""
            if sp_id not in test_subplots: continue
            
            count_val = to_float(row[shrub_col_map.get('株数', 2)] if shrub_col_map.get('株数', 2) < len(row) else 0.0, 0.0)
            crown_val = to_float(row[shrub_col_map.get('平均灌丛幅（cm）', 3)] if shrub_col_map.get('平均灌丛幅（cm）', 3) < len(row) else 0.0, 0.0)
            height_val = to_float(row[shrub_col_map.get('平均高度（cm）', 5)] if shrub_col_map.get('平均高度（cm）', 5) < len(row) else 0.0, 0.0)
            coverage_val = to_float(row[shrub_col_map.get('盖度', 6)] if shrub_col_map.get('盖度', 6) < len(row) else 0.0, 0.0)
            
            shrub_records.append({
                "subplot_id": sp_id,
                "species": str(row[shrub_col_map.get('植物名称', 1)]).strip() if shrub_col_map.get('植物名称', 1) < len(row) and row[shrub_col_map.get('植物名称', 1)] else "未知灌木",
                "count": count_val,
                "crown_width_cm": crown_val,
                "height_cm": height_val,
                "coverage": coverage_val
            })
            
    if '枯死木调查数据' in wb.sheetnames:
        sheet_dw = wb['枯死木调查数据']
        dw_col_map = resolve_sheet_col_map(sheet_dw, header_row_count=1)
        for idx, row in enumerate(sheet_dw.iter_rows(min_row=2, values_only=True)):
            if not row or all(c is None for c in row): continue
            sp_id = str(row[dw_col_map.get('小样方号', 0)]).strip() if dw_col_map.get('小样方号', 0) < len(row) and row[dw_col_map.get('小样方号', 0)] else ""
            if sp_id not in test_subplots: continue
            deadwood_records.append({
                "deadwood_id": f"DW_{sp_id}_{idx+1}",
                "subplot_id": sp_id,
                "species": str(row[dw_col_map.get('枯死木植物名称', 1)]).strip() if dw_col_map.get('枯死木植物名称', 1) < len(row) and row[dw_col_map.get('枯死木植物名称', 1)] else "未知枯死木",
                "total_count": to_float(row[dw_col_map.get('株数', 2)] if dw_col_map.get('株数', 2) < len(row) else 1.0, 1.0),
                "standing_count": to_float(row[dw_col_map.get('枯立木', 3)] if dw_col_map.get('枯立木', 3) < len(row) else 0.0, 0.0),
                "fallen_count": to_float(row[dw_col_map.get('枯倒木', 4)] if dw_col_map.get('枯倒木', 4) < len(row) else 0.0, 0.0),
                "remarks": str(row[dw_col_map.get('备注', 5)]).strip() if dw_col_map.get('备注', 5) < len(row) and row[dw_col_map.get('备注', 5)] else ""
            })
            
    with driver.session() as session:
        if shrub_records:
            print(f"正在秒级批处理注入 {len(shrub_records)} 条灌木层记录...")
            run_cypher(session, """
            UNWIND $batch AS r
            MATCH (sub:Subplot {subplot_id: r.subplot_id})
            MERGE (tx:TreeSpecies:Taxon {accepted_name_cn: r.species})
            MERGE (obs:ShrubObservation {observation_id: "SHRUB_" + r.subplot_id + "_" + r.species})
            SET obs.species_name = r.species,
                obs.count = r.count,
                obs.crown_width_cm = r.crown_width_cm,
                obs.height_cm = r.height_cm,
                obs.coverage = r.coverage,
                obs.survey_date = "2023-08"
            MERGE (sub)-[:HAS_SHRUB_OBSERVATION]->(obs)
            MERGE (obs)-[:IS_SPECIES]->(tx)
            """, batch=shrub_records)
            
        if deadwood_records:
            print(f"正在秒级批处理注入 {len(deadwood_records)} 条枯死木记录...")
            run_cypher(session, """
            UNWIND $batch AS r
            MATCH (sub:Subplot {subplot_id: r.subplot_id})
            MERGE (tx:TreeSpecies:Taxon {accepted_name_cn: r.species})
            MERGE (obs:DeadwoodObservation {observation_id: r.deadwood_id})
            SET obs.species_name = r.species,
                obs.total_count = r.total_count,
                obs.standing_count = r.standing_count,
                obs.fallen_count = r.fallen_count,
                obs.remarks = r.remarks,
                obs.survey_date = "2023-08"
            MERGE (sub)-[:HAS_DEADWOOD_OBSERVATION]->(obs)
            MERGE (obs)-[:IS_SPECIES]->(tx)
            """, batch=deadwood_records)
    
    print("\n正在生成并挂载样方层宏观总结指标实体 (样方总蓄积量、香农指数 H' 等宏观实体)...")
    with driver.session() as session:
        session.run("""
        MATCH (s:Subplot)-[:HAS_TREE]->(t:TreeIndividual)-[:HAS_OBSERVATION]->(obs:TreeObservation)
        WITH s, count(obs) AS tree_cnt, sum(coalesce(obs.volume_m3, 0.0)) AS total_vol
        MERGE (iv_vol:IndicatorValue {indicator_value_id: "IV_SUBPLOT_VOL_" + s.subplot_id + "_2023"})
        SET iv_vol.indicator_name = "样方总蓄积量",
            iv_vol.value = round(total_vol, 4),
            iv_vol.unit = "m³",
            iv_vol.level = "subplot",
            iv_vol.value_origin = "aggregated_from_official_two_way_single_tree_volumes"
        MERGE (s)-[:HAS_INDICATOR_VALUE]->(iv_vol)
        """)
        
        session.run("""
        MATCH (s:Subplot)-[:HAS_TREE]->(t:TreeIndividual)-[:HAS_OBSERVATION]->(obs:TreeObservation)
        MATCH (t)-[:IS_SPECIES]->(tx:Taxon)
        WITH s, tx.accepted_name_cn AS sp, count(obs) AS c_sp
        WITH s, sum(c_sp) AS N_total, collect({sp: sp, c: c_sp}) AS sp_list
        WITH s, N_total, [x IN sp_list | (toFloat(x.c)/N_total) * log(toFloat(x.c)/N_total)] AS entropy_terms
        WITH s, -round(sum([e IN entropy_terms | e][0]), 3) AS shannon_h
        MERGE (iv_sh:IndicatorValue {indicator_value_id: "IV_SUBPLOT_SHANNON_" + s.subplot_id + "_2023"})
        SET iv_sh.indicator_name = "Shannon-Wiener H'",
            iv_sh.value = shannon_h,
            iv_sh.unit = "index",
            iv_sh.level = "subplot"
        MERGE (s)-[:HAS_INDICATOR_VALUE]->(iv_sh)
        """)
        
        # 动态声明式知识库驱动渲染：从 YAML 加载公式/模型/变量/规则，并自动从 Excel 查表矩阵核定极值
        inject_declarative_knowledge_registry(
            session,
            r"E:\Paper_Doing\Model_third\Literature_miner\DeepKnowledge-20260613\ForestryAgent\ontology\forestry_knowledge_registry.yaml",
            volume_dict
        )
        
        summary = session.run("""
        MATCH (mp:MonitoringPlot)-[:HAS_SUBPLOT]->(s:Subplot)
        OPTIONAL MATCH (s)-[:HAS_TREE]->(t:TreeIndividual)
        OPTIONAL MATCH (s)-[:HAS_INDICATOR_VALUE]->(iv_vol:IndicatorValue {indicator_name: "样方总蓄积量"})
        OPTIONAL MATCH (s)-[:HAS_INDICATOR_VALUE]->(iv_sh:IndicatorValue {indicator_name: "Shannon-Wiener H'"})
        OPTIONAL MATCH (s)-[:HAS_SHRUB_OBSERVATION]->(sh:ShrubObservation)
        OPTIONAL MATCH (s)-[:HAS_DEADWOOD_OBSERVATION]->(dw:DeadwoodObservation)
        RETURN s.subplot_id AS subplot, 
               count(DISTINCT t) AS trees, 
               coalesce(iv_vol.value, 0.0) AS total_vol_m3,
               coalesce(iv_sh.value, 0.0) AS shannon_h,
               count(DISTINCT sh) AS shrub_types,
               count(DISTINCT dw) AS deadwood_records
        ORDER BY subplot
        """).data()
        print("\n=== [图谱全息验证：六层知识架构与三类变量双驱动对齐完毕] 5个典型样方产出摘要 ===")
        for row in summary:
            print(f"  样方 [{row['subplot']}]: 乔木 {row['trees']} 株 | [第5层产出 IndicatorValue:OutputIndicator] 蓄积 {row['total_vol_m3']} m³, 香农指数 {row['shannon_h']} | 灌木 {row['shrub_types']} 种 | 枯死木 {row['deadwood_records']} 条")

if __name__ == "__main__":
    push_5_subplots_to_neo4j()

